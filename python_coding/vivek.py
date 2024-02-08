import openpyxl
import pandas as pd 


def process_rows(filename, target_file_name):
    scopes = ["Global", "Folder scope", "File scope"]
    result = []
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    second_row_values = [cell for cell in sheet.iter_rows(min_row=2, max_row=2, values_only=True)][0]
    count = 0
    current_row_values = []
    if second_row_values[1] != "Global":
        for row in sheet.iter_rows(min_row=1):  # Include the header row     
            if row[0].value is not None and count == 1:
                return result
            if row[0].value == target_file_name:
                current_row_values = [cell.value.strip() for cell in row[1:] if cell.value]  # Accumulate values
                if any(current_row_values):  # Check if any values are present
                    output = "_".join(current_row_values)
                    result.append(output)
                    count = 1    
            elif row[0].value is None and count == 1:
                current_row_values = [cell.value.strip() for cell in row[1:] if cell.value]  # Accumulate values
                if any(current_row_values):  # Check if any values are present
                    output = "_".join(current_row_values)
                    result.append(output)
        return result     
    else:
        state = ""
        type2_result = {scopes[0]: [], scopes[1]: [], scopes[2]: []}
        stage_two = False
        for row in sheet.iter_rows(min_row=1, values_only=True): 
            if row[0] is not None and stage_two == True:
                break
            if row[0] == target_file_name:
                stage_two = True
                state = scopes[0]
                current_row_values = [cell.strip() for cell in row[2:] if cell]
                output = "_".join(current_row_values)
                type2_result[state].append(output)
            elif row[0] is None and row[1] is None and stage_two == True:
                current_row_values = [cell.strip() for cell in row[2:] if cell]
                output = "_".join(current_row_values)
                type2_result[state].append(output)
            elif row[1] in scopes and stage_two == True:
                state = row[1]
                current_row_values = [cell.strip() for cell in row[2:] if cell]
                output = "_".join(current_row_values)
                type2_result[state].append(output)  

    print(type2_result, '=============type3333333333')

if __name__ == "__main__":
    filename = r"Book1.xlsx"  # Replace with your actual file name
    target_file_name = "Swc_Observer"
    x = process_rows(filename, target_file_name)
    print(x)
# Swc_Observer